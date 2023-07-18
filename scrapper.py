import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime

def stores_data(wb):
    # Fetch the website data
    url = 'https://hotpoint.co.ke/stores/?gclid=Cj0KCQjwzdOlBhCNARIsAPMwjbwxLfqdX49U2hmPC5229iaGfASk6-qzf_3nH68UzPc4mVPRWi1zAlcaAg0tEALw_wcB'
    response = requests.get(url)
    html_content = response.content

    # Parse the HTML and extract data
    soup = BeautifulSoup(html_content, 'html.parser')
    store_details = []
    store_elements = soup.find_all('div', class_='stores-store-body')
    for store_element in store_elements:
        name = store_element.find('h3').text.strip()
        location = store_element.find('address').text.strip()
        tel = store_element.find('p', class_='mb-0').text.strip()
        email = store_element.find('a', class_='email').text.strip()
        store_details.append((name, location, tel, email))

    # Create a new sheet or select the "stores" sheet in the existing Excel file
    sheet = wb.create_sheet(title="stores", index=0)

    # Write the headers
    sheet['A1'] = "Store Name"
    sheet['B1'] = "Location"
    sheet['C1'] = "Telephone"
    sheet['D1'] = "Email"

    # Write the store details to the Excel sheet
    for i, details in enumerate(store_details, start=2):
        sheet.cell(row=i, column=1).value = details[0]  # Store Name
        sheet.cell(row=i, column=2).value = details[1]  # Location
        sheet.cell(row=i, column=3).value = details[2]  # Telephone
        sheet.cell(row=i, column=4).value = details[3]  # Email

def categories_data(wb):
    # Fetch the website data for categories
    url = 'https://hotpoint.co.ke/categories'
    response = requests.get(url)
    html_content = response.content

    # Parse the HTML and extract data for categories
    soup = BeautifulSoup(html_content, 'html.parser')
    category_details = []
    category_elements = soup.find_all('a', class_='nav-link dropdown-toggle')
    for category_element in category_elements:
        href_value = category_element['href']
        text_value = category_element.text.strip()
        category_details.append((href_value, text_value))

    # Create a new sheet or select the "categories" sheet in the existing Excel file
    sheet = wb.create_sheet(title="categories", index=1)

    # Write the headers for categories
    sheet['A1'] = "Category"
    sheet['B1'] = "Category Text"

    # Write the category details to the Excel sheet
    for i, details in enumerate(category_details, start=2):
        sheet.cell(row=i, column=1).value = details[0]  # Category
        sheet.cell(row=i, column=2).value = details[1]  # Category Text

from bs4 import BeautifulSoup

from bs4 import BeautifulSoup

def tv_data(wb):
    # Fetch the website data for categories
    url = 'https://hotpoint.co.ke/catalogue/category/tv-entertainment/'
    response = requests.get(url)
    html_content = response.content

    # Parse the HTML and extract data for categories
    soup = BeautifulSoup(html_content, 'html.parser')
    category_details = []
    category_elements = soup.find_all('li')
    for category_element in category_elements:
        category_link = category_element.find('a')['href']
        category_text = category_element.find('a').text.strip()
        category_details.append((category_link, category_text))

    # Create a new sheet or select the "tv_entertainment" sheet in the existing Excel file
    sheet = wb.create_sheet(title="tv_entertainment", index=2)

    # Write the headers for the columns
    sheet['A1'] = "Category"
    sheet['B1'] = "Products"

    # Write the category details to the Excel sheet
    for i, detail(row=i, column=1).value = details[0]  # Category (href value)
        sheet.cellls in enumerate(category_details, start=2):
        sheet.cel(row=i, column=2).value = details[1]  # Products (text value)

# Create an Excel workbook
wb = Workbook()

# Call the functions to scrape and save the data
stores_data(wb)
categories_data(wb)
tv_data(wb)

# Get the current date and time
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Generate the file name with the current date and time
file_name = f"hotpoint_data_{current_datetime}.xlsx"

# Specify the path to the "logs" folder
folder_path = os.path.join(os.getcwd(), "logs")

# Create the "logs" folder if it doesn't exist
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# Save the Excel file in the "logs" folder
file_path = os.path.join(folder_path, file_name)
wb.save(file_path)
